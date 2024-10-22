import time
import openpyxl
import requests
from bs4 import BeautifulSoup
import re
import streamlit as st
import os
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome import service as fs
from selenium.webdriver import ChromeOptions
from webdriver_manager.core.os_manager import ChromeType
from selenium.webdriver.common.by import By

# スクレイピングしたいURL（例: 福岡の求人）
url = 'https://www.staff-q.co.jp/recruiter'

# エクセルを開く
wb = openpyxl.Workbook()
ws = wb.active

st.title('マイナビシニアの求人検索')

# 初期の値を表示するための空の場所を作成
text_placeholder = st.empty()

# 初期化
if 'stop' not in st.session_state:
    st.session_state.stop = False
if 'processing' not in st.session_state:
    st.session_state.processing = False


if st.button('開始', disabled=st.session_state.processing):
    st.session_state.stop = False
    st.text('ルート確認２')
    if url:
        # ドライバのオプション
        options = ChromeOptions()

        # option設定を追加（設定する理由はメモリの削減）
        options.add_argument("--headless")
        options.add_argument('--disable-gpu')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')

        # webdriver_managerによりドライバーをインストール
        CHROMEDRIVER = ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install()
        service = fs.Service(CHROMEDRIVER)
        driver = webdriver.Chrome(
                                options=options,
                                service=service
                                )

        # URLで指定したwebページを開く
        driver.get(url)
        c = 1  # 行数
        j = 1  # ページ数
        while c <= 10:
            print(f'Processing page {j}...')
            soup = BeautifulSoup(response.text, 'html.parser')
            job_cards = soup.find_all('div', class_='job-summary__header')
            table_cards = soup.find_all('table', class_='job-summary-table')

            if not job_cards:
                print('求人が見つかりませんでした。')
                break
            for job, job_table in zip(job_cards, table_cards):
                print('testcount:', c)

                if st.session_state.stop:
                    st.session_state.running = False
                    st.warning("処理が中断されました。")
                    break

                # プレースホルダーに新しいテキストを表示
                text_placeholder.text(f'読込数：{c}')

                # 求人タイトルを取得
                title = job.find('a', target='_blank').text.strip()
                print('求人タイトル:', title)
                
                # 会社名を取得
                company = job.find('p', class_='job-summary__name').text.strip()
                print('会社名:', company)

                # テーブルデータを取得（例として最初のテーブルを取得）
                rows = job_table.find_all('tr')

                # テーブルデータをリストに格納
                table_data = []
                for row in rows:
                    cols = row.find_all('td')
                    cols = [col.get_text(strip=True) for col in cols]
                    if cols:  # 空でない場合のみ追加
                        table_data.append(cols)

                # 1つのセルにデータを書き込むための文字列を作成
                cell_data = '\n'.join(['\t'.join(row) for row in table_data])


                # 詳細ページのリンクを取得
                job_link = 'https://mynavi-ms.jp/' + job.find('a')['href']

                # 待機
                time.sleep(10)
                # 詳細ページにリクエストを送信
                job_response = requests.get(job_link)
                job_soup = BeautifulSoup(job_response.text, 'html.parser')

                # 電話番号を正規表現で抽出（例: 012-345-6789）
                phone_numbers = re.findall(r'\d{2,4}-\d{2,4}-\d{2,4}', job_soup.text)
                phone_numbers2 = [number for number in phone_numbers if not number.startswith("0120")]
                if phone_numbers2:
                    print('電話番号:', phone_numbers)
                    # Excelファイルに書き込mi 
                    ws.cell(row=c, column=1, value=title)
                    ws.cell(row=c, column=2, value=company)
                    ws.cell(row=c, column=3, value=', '.join(phone_numbers) if phone_numbers else 'なし')
                    ws.cell(row=c, column=4, value=cell_data)
                    c += 1
                else:
                    print('電話番号は見つかりませんでした。')
                
                print('-' * 40)
            # 次のページを探す
            next_page = soup.find('a', class_='search-pagenation__next')
            if next_page and 'href' in next_page.attrs:
                next_url = 'https://mynavi-ms.jp' + next_page['href']
                driver.get(next_url)
                time.sleep(5)  # サーバーへの負荷を避けるためにスリープ
                j += 1
            else:
                print('pageEnd')
                break  # 次のページがない場合は終了
        else:
            st.text('完了しました。')
            #st.text("C:/aggregate/に保存されました。")

# エクセルファイルを保存
wb.save("mynavi-ms.xlsx")
wb.close()
st.session_state.processing = False

if st.button("中断", disabled=st.session_state.processing):
    st.session_state.stop = True